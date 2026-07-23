# app/data.py
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

DEFAULT_DATASET_PATH = (
    Path(__file__).resolve().parent.parent / "data" / "Bluno_TradeFinanceAgent_Dataset_v1.xlsx"
)


def _rows(ws):
    """Yield sheet rows as dicts keyed by the header row."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = [str(h).strip() for h in rows[0]]
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        yield dict(zip(header, row))


class Database:
    def __init__(self, path: str | Path):
        wb = load_workbook(path, data_only=True, read_only=True)
        companies = list(_rows(wb["Companies"]))
        sanctions = list(_rows(wb["Sanctions"]))
        credit = list(_rows(wb["Credit_Limits"]))
        exposure = list(_rows(wb["Exposures"]))
        shipments = list(_rows(wb["Shipments"]))
        wb.close()

        self._companies = companies
        self._companies_by_name = {c["company_name"].strip().lower(): c for c in companies}
        self._clearance_by_country = {s["country"].strip().lower(): s for s in sanctions}
        self._credit_by_id = {c["company_id"]: c["credit_limit_usd"] for c in credit}
        self._exposure_by_id = {e["company_id"]: e["outstanding_exposure_usd"] for e in exposure}
        self._shipments = shipments

    def resolve_company(self, name: str) -> dict | None:
        c = self._companies_by_name.get(name.strip().lower())
        if c is None:
            return None
        return {
            "company_id": c["company_id"],
            "company_name": c["company_name"],
            "destination_country": c["destination_country"],
            "kyc_status": c["kyc_status"],
        }

    def get_clearance(self, country: str) -> dict | None:
        s = self._clearance_by_country.get(country.strip().lower())
        if s is None:
            return None
        return {"country": s["country"], "clearance": s["clearance"], "note": s["note"]}

    def get_credit(self, name: str) -> dict | None:
        c = self.resolve_company(name)
        if c is None:
            return None
        cid = c["company_id"]
        limit = self._credit_by_id.get(cid)
        exposure = self._exposure_by_id.get(cid)
        if limit is None or exposure is None:
            return None
        return {
            "company_id": cid,
            "company_name": c["company_name"],
            "credit_limit_usd": limit,
            "outstanding_exposure_usd": exposure,
            "available_headroom_usd": limit - exposure,
        }

    def get_shipments(self, name: str) -> dict | None:
        c = self.resolve_company(name)
        if c is None:
            return None
        cid = c["company_id"]
        ships = [
            {
                "shipment_id": s["shipment_id"],
                "counterparty_name": s["counterparty_name"],
                "value_usd": s["value_usd"],
                "status": s["status"],
            }
            for s in self._shipments
            if s["company_id"] == cid
        ]
        return {"company_id": cid, "company_name": c["company_name"], "shipments": ships}

    def filter_companies(
        self, destination_country: str | None = None, kyc_status: str | None = None
    ) -> list[dict]:
        out = []
        for c in self._companies:
            if destination_country is not None and (
                c["destination_country"].strip().lower() != destination_country.strip().lower()
            ):
                continue
            if kyc_status is not None and (
                str(c["kyc_status"]).strip().lower() != kyc_status.strip().lower()
            ):
                continue
            out.append(
                {
                    "company_id": c["company_id"],
                    "company_name": c["company_name"],
                    "destination_country": c["destination_country"],
                    "kyc_status": c["kyc_status"],
                }
            )
        return out


_DB: Database | None = None


def init_db(path: str | Path = DEFAULT_DATASET_PATH) -> Database:
    global _DB
    _DB = Database(path)
    return _DB


def get_db() -> Database:
    global _DB
    if _DB is None:
        _DB = Database(DEFAULT_DATASET_PATH)
    return _DB
